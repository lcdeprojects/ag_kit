import openpyxl
from openpyxl.styles import Font, PatternFill
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.db import models
from django.utils import timezone
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from .decorator import group_required
from django.contrib.auth.models import User
from .models import Patient, Professional, HealthPlan, Payment, Appointment

from django import forms

class CrudMixin:
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['verbose_name'] = self.model._meta.verbose_name
        context['verbose_name_plural'] = self.model._meta.verbose_name_plural
        return context

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        for field_name, field in form.fields.items():
            # Add Bootstrap classes
            if isinstance(field.widget, (forms.Select, forms.SelectMultiple)):
                field.widget.attrs.update({'class': 'form-select'})
            else:
                field.widget.attrs.update({'class': 'form-control'})
            
            # Add HTML5 date/time pickers for Date and DateTime fields
            try:
                model_field = self.model._meta.get_field(field_name)
                if isinstance(model_field, models.DateTimeField):
                    field.widget = forms.DateTimeInput(
                        attrs={'class': 'form-control', 'type': 'datetime-local'},
                        format='%Y-%m-%dT%H:%M'
                    )
                elif isinstance(model_field, models.DateField):
                    field.widget = forms.DateInput(
                        attrs={'class': 'form-control', 'type': 'date'},
                        format='%Y-%m-%d'
                    )
            except models.FieldDoesNotExist:
                pass
        return form

@login_required
def dashboard(request):
    today = timezone.now().date()
    start_of_month = today.replace(day=1)
    
    # Filter for expiring or expired plans
    all_payments = Payment.objects.all().order_by('payment_date')
    expiring_plans = [p for p in all_payments if p.status != 'active']
    
    # Separate counts
    expired_count = len([p for p in expiring_plans if p.status == 'expired'])
    warning_count = len([p for p in expiring_plans if p.status == 'warning'])
    
    # Permissions check
    is_admin = request.user.is_superuser or request.user.groups.filter(name='Administradores').exists()
    
    # Calculate revenue only for admins
    month_revenue = 0
    if is_admin:
        month_revenue = Payment.objects.filter(payment_date__gte=start_of_month).aggregate(models.Sum('amount'))['amount__sum'] or 0

    context = {
        'total_patients': Patient.objects.count(),
        'total_professionals': Professional.objects.count(),
        'today_appointments': Appointment.objects.filter(date__date=today).count(),
        'month_revenue': month_revenue,
        'upcoming_appointments': Appointment.objects.filter(date__date=today).order_by('date')[:10],
        'expiring_plans': expiring_plans[:5],
        'expired_count': expired_count,
        'warning_count': warning_count,
        'is_admin': is_admin,
    }
    return render(request, 'clinic/dashboard.html', context)

# Patient CRUD
@group_required('Administradores','Profissionais','Financeiro')
class PatientListView(LoginRequiredMixin, ListView):
    model = Patient
    template_name = 'clinic/patient_list.html'
    context_object_name = 'patients'
    paginate_by = 6

    def get_queryset(self):
        qs = Patient.objects.all().order_by('first_name', 'last_name')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(models.Q(first_name__icontains=q) | models.Q(last_name__icontains=q))
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_q'] = self.request.GET.get('q', '')
        return context

@group_required('Administradores','Profissionais','Financeiro')
class PatientCreateView(LoginRequiredMixin, CrudMixin, CreateView):
    model = Patient
    fields = ['first_name', 'last_name', 'cpf', 'birth_date', 'phone', 'email']
    template_name = 'clinic/generic_form.html'
    success_url = reverse_lazy('patient-list')

@group_required('Administradores','Profissionais','Financeiro')
class PatientUpdateView(LoginRequiredMixin, CrudMixin, UpdateView):
    model = Patient
    fields = ['first_name', 'last_name', 'cpf', 'birth_date', 'phone', 'email']
    template_name = 'clinic/generic_form.html'
    success_url = reverse_lazy('patient-list')

@group_required('Administradores','Profissionais','Financeiro')
class PatientDeleteView(LoginRequiredMixin, CrudMixin, DeleteView):
    model = Patient
    template_name = 'clinic/generic_confirm_delete.html'
    success_url = reverse_lazy('patient-list')

# Professional CRUD
@group_required('Administradores')
class ProfessionalListView(LoginRequiredMixin, ListView):
    model = Professional
    template_name = 'clinic/professional_list.html'
    context_object_name = 'professionals'

@group_required('Administradores')
class ProfessionalCreateView(LoginRequiredMixin, CrudMixin, CreateView):
    model = Professional
    fields = ['name', 'registration_number', 'specialty', 'professional_type']
    template_name = 'clinic/generic_form.html'
    success_url = reverse_lazy('professional-list')

@group_required('Administradores')
class ProfessionalUpdateView(LoginRequiredMixin, CrudMixin, UpdateView):
    model = Professional
    fields = ['name', 'registration_number', 'specialty', 'professional_type']
    template_name = 'clinic/generic_form.html'
    success_url = reverse_lazy('professional-list')

@group_required('Administradores')
class ProfessionalDeleteView(LoginRequiredMixin, CrudMixin, DeleteView):
    model = Professional
    template_name = 'clinic/generic_confirm_delete.html'
    success_url = reverse_lazy('professional-list')

# HealthPlan CRUD
@group_required('Administradores')
class HealthPlanListView(LoginRequiredMixin, ListView):
    model = HealthPlan
    template_name = 'clinic/healthplan_list.html'
    context_object_name = 'plans'

@group_required('Administradores')
class HealthPlanCreateView(LoginRequiredMixin, CrudMixin, CreateView):
    model = HealthPlan
    fields = ['title', 'description', 'validity_days', 'amount', 'percentage_medical_plan', 'percentage_nutrition_plan']
    template_name = 'clinic/generic_form.html'
    success_url = reverse_lazy('healthplan-list')

@group_required('Administradores')
class HealthPlanUpdateView(LoginRequiredMixin, CrudMixin, UpdateView):
    model = HealthPlan
    fields = ['title', 'description', 'validity_days', 'amount', 'percentage_medical_plan', 'percentage_nutrition_plan']
    template_name = 'clinic/generic_form.html'
    success_url = reverse_lazy('healthplan-list')

@group_required('Administradores')
class HealthPlanDeleteView(LoginRequiredMixin, CrudMixin, DeleteView):
    model = HealthPlan
    template_name = 'clinic/generic_confirm_delete.html'
    success_url = reverse_lazy('healthplan-list')

# Payment CRUD
@group_required('Administradores','Financeiro')
class PaymentListView(LoginRequiredMixin, ListView):
    model = Payment
    template_name = 'clinic/payment_list.html'
    context_object_name = 'payments'
    paginate_by = 4

    def get_queryset(self):
        qs = Payment.objects.all().order_by('-payment_date')
        
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(models.Q(patient__first_name__icontains=q) | models.Q(patient__last_name__icontains=q))
            
        plan_id = self.request.GET.get('plan')
        if plan_id:
            qs = qs.filter(plan_id=plan_id)
            
        start_date = self.request.GET.get('start')
        if start_date:
            qs = qs.filter(payment_date__gte=start_date)
            
        end_date = self.request.GET.get('end')
        if end_date:
            qs = qs.filter(payment_date__lte=end_date)
            
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['plans'] = HealthPlan.objects.all()
        context['current_q'] = self.request.GET.get('q', '')
        context['current_plan'] = self.request.GET.get('plan', '')
        context['current_start'] = self.request.GET.get('start', '')
        context['current_end'] = self.request.GET.get('end', '')
        return context

@group_required('Administradores','Financeiro')
class PaymentCreateView(LoginRequiredMixin, CrudMixin, CreateView):
    model = Payment
    fields = ['patient', 'plan', 'amount', 'payment_date', 'notes']
    template_name = 'clinic/payment_form.html'
    success_url = reverse_lazy('payment-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pass plan amounts for JS automation
        plans = HealthPlan.objects.all()
        context['plan_amounts'] = {plan.id: float(plan.amount) for plan in plans}
        return context

@group_required('Administradores','Financeiro')
class PaymentUpdateView(LoginRequiredMixin, CrudMixin, UpdateView):
    model = Payment
    fields = ['patient', 'plan', 'amount', 'payment_date', 'notes']
    template_name = 'clinic/payment_form.html'
    success_url = reverse_lazy('payment-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        plans = HealthPlan.objects.all()
        context['plan_amounts'] = {plan.id: float(plan.amount) for plan in plans}
        return context

@group_required('Administradores','Financeiro')
class PaymentDeleteView(LoginRequiredMixin, CrudMixin, DeleteView):
    model = Payment
    template_name = 'clinic/generic_confirm_delete.html'
    success_url = reverse_lazy('payment-list')

# Appointment CRUD

@group_required('Administradores','Profissionais')
class AppointmentListView(LoginRequiredMixin, ListView):
    model = Appointment
    template_name = 'clinic/appointment_list.html'
    context_object_name = 'appointments'
    paginate_by = 6

    def get_queryset(self):
        qs = Appointment.objects.all().order_by('-date')
        
        # Pesquisa por nome do paciente
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(models.Q(patient__first_name__icontains=q) | models.Q(patient__last_name__icontains=q))
            
        # Filtro por data
        date_filter = self.request.GET.get('date')
        if date_filter:
            qs = qs.filter(date__date=date_filter)
            
        # Filtro por responsável
        user_filter = self.request.GET.get('user')
        if user_filter:
            qs = qs.filter(user_id=user_filter)
            
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['users'] = User.objects.all().order_by('first_name', 'username')
        context['current_q'] = self.request.GET.get('q', '')
        context['current_date'] = self.request.GET.get('date', '')
        context['current_user'] = self.request.GET.get('user', '')
        return context

@group_required('Administradores','Profissionais')
class AppointmentCreateView(LoginRequiredMixin, CrudMixin, CreateView):
    model = Appointment
    fields = ['patient', 'date', 'status', 'weight', 'clinical_notes', 'prescription']
    template_name = 'clinic/generic_form.html'
    success_url = reverse_lazy('appointment-list')

    def get_initial(self):
        initial = super().get_initial()
        patient_id = self.request.GET.get('patient')
        if patient_id:
            initial['patient'] = patient_id
        return initial

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)

@group_required('Administradores','Profissionais')
class AppointmentUpdateView(LoginRequiredMixin, CrudMixin, UpdateView):
    model = Appointment
    fields = ['patient', 'date', 'status', 'weight', 'clinical_notes', 'prescription']
    template_name = 'clinic/generic_form.html'
    success_url = reverse_lazy('appointment-list')

@group_required('Administradores','Profissionais')
class AppointmentDetailView(LoginRequiredMixin, CrudMixin, DeleteView): # Inheriting from DeleteView just for CrudMixin context, better use DetailView
    model = Appointment
    template_name = 'clinic/appointment_detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['verbose_name'] = 'Detalhes da Consulta'
        return context

@group_required('Administradores','Profissionais')
class PatientHistoryView(LoginRequiredMixin, ListView):
    model = Appointment
    template_name = 'clinic/patient_history.html'
    context_object_name = 'appointments'

    def get_queryset(self):
        return Appointment.objects.filter(patient_id=self.kwargs['pk']).order_by('-date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        patient = Patient.objects.get(pk=self.kwargs['pk'])
        context['patient'] = patient
        
        # Data for Weight Chart
        weight_data = Appointment.objects.filter(
            patient=patient, 
            weight__isnull=False
        ).order_by('date')
        
        context['weight_labels'] = [a.date.strftime("%d/%m/%Y") for a in weight_data]
        context['weight_values'] = [float(a.weight) for a in weight_data]
        
        return context

@group_required('Administradores','Profissionais')
class AppointmentDeleteView(LoginRequiredMixin, CrudMixin, DeleteView):
    model = Appointment
    template_name = 'clinic/generic_confirm_delete.html'
    success_url = reverse_lazy('appointment-list')

class AgendaView(LoginRequiredMixin, CrudMixin, ListView):
    # Using ListView to inherit CrudMixin context, even though we won't list a specific model here
    # We'll override the model name in the template
    model = Appointment 
    template_name = 'clinic/agenda.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['verbose_name'] = 'Agenda'
        return context

@group_required('Administradores','Financeiro')
def export_payment_report(request):
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    
    payments = Payment.objects.all()
    if start_date:
        payments = payments.filter(payment_date__gte=start_date)
    if end_date:
        payments = payments.filter(payment_date__lte=end_date)
        
    # Excel setup
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Pagamentos"
    
    # Headers
    headers = ["ID", "Paciente", "Plano", "Valor", "Data", "Data de Expiração", "Dias Restantes", "Status", "Valor Médico", "Valor Nutricionista"]
    ws.append(headers)
    
    # Header Styling
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    
    # Data
    for p in payments:
        ws.append([
            p.id,
            p.patient.name,
            p.plan.title if p.plan else "N/A",
            p.amount,
            p.payment_date.strftime("%d/%m/%Y"),
            p.expiration_date.strftime("%d/%m/%Y") if p.expiration_date else "N/A",
            p.days_remaining,
            p.status,
            p.amount_medical_plan,  
            p.amount_nutrition_plan,
        ])
    # Calculate Totals
    total_amount = sum(float(p.amount) for p in payments)
    total_medical = sum(float(p.amount_medical_plan) for p in payments)
    total_nutrition = sum(float(p.amount_nutrition_plan) for p in payments)
    
    # Summary Row
    summary_row = ["TOTAL", "", "", total_amount, "", "", "", "", total_medical, total_nutrition]
    ws.append(summary_row)
    
    # Summary Styling
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    # Response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename=relatorio_pagamentos_{timezone.now().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response

def denied(request):
    return render(request, 'clinic/denied.html')


def today(request):
    appointments_today = Appointment.objects.filter(date__date=timezone.now().date()).order_by('date')
    return render(request, 'clinic/today.html', {'appointments': appointments_today})
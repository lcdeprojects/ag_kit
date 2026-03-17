import openpyxl
from openpyxl.styles import Font, PatternFill
from django.shortcuts import redirect
from django.http import HttpResponse
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import models
from django.utils import timezone
from .base import CrudMixin
from ..decorator import group_required
from ..models import HealthPlan, Payment

@group_required('Administradores','Financeiro')
class PaymentListView(LoginRequiredMixin, ListView):
    model = Payment
    template_name = 'clinic/payment_list.html'
    context_object_name = 'payments'
    paginate_by = 4

    def get_queryset(self):
        qs = Payment.objects.all().order_by('-payment_date')
        
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(models.Q(patient__first_name__icontains=q) | models.Q(patient__last_name__icontains=q))
            
        plan_id = self.request.GET.get('plan')
        if plan_id:
            qs = qs.filter(plan_id=plan_id)
            
        start_date = self.request.GET.get('start')
        if start_date:
            qs = qs.filter(payment_date__gte=start_date)
            
        end_date = self.request.GET.get('end')
        if end_date:
            qs = qs.filter(payment_date__lte=end_date)
            
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['plans'] = HealthPlan.objects.all()
        context['current_q'] = self.request.GET.get('q', '')
        context['current_plan'] = self.request.GET.get('plan', '')
        context['current_start'] = self.request.GET.get('start', '')
        context['current_end'] = self.request.GET.get('end', '')
        return context

@group_required('Administradores','Financeiro')
class PaymentCreateView(LoginRequiredMixin, CrudMixin, CreateView):
    model = Payment
    fields = ['patient', 'plan', 'amount', 'payment_date', 'notes']
    template_name = 'clinic/payment_form.html'
    success_url = reverse_lazy('payment-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pass plan amounts for JS automation
        plans = HealthPlan.objects.all()
        context['plan_amounts'] = {plan.id: float(plan.amount) for plan in plans}
        return context

@group_required('Administradores','Financeiro')
class PaymentUpdateView(LoginRequiredMixin, CrudMixin, UpdateView):
    model = Payment
    fields = ['patient', 'plan', 'amount', 'payment_date', 'notes']
    template_name = 'clinic/payment_form.html'
    success_url = reverse_lazy('payment-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        plans = HealthPlan.objects.all()
        context['plan_amounts'] = {plan.id: float(plan.amount) for plan in plans}
        return context

@group_required('Administradores','Financeiro')
class PaymentDeleteView(LoginRequiredMixin, CrudMixin, DeleteView):
    model = Payment
    template_name = 'clinic/generic_confirm_delete.html'
    success_url = reverse_lazy('payment-list')

@group_required('Administradores','Financeiro')
def export_payment_report(request):
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    
    payments = Payment.objects.all()
    if start_date:
        payments = payments.filter(payment_date__gte=start_date)
    if end_date:
        payments = payments.filter(payment_date__lte=end_date)
        
    # Excel setup
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Pagamentos"
    
    # Headers
    headers = ["ID", "Paciente", "Plano", "Valor", "Data", "Data de Expiração", "Dias Restantes", "Status", "Valor Médico", "Valor Nutricionista"]
    ws.append(headers)
    
    # Header Styling
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    
    # Data
    for p in payments:
        ws.append([
            p.id,
            p.patient.name,
            p.plan.title if p.plan else "N/A",
            p.amount,
            p.payment_date.strftime("%d/%m/%Y"),
            p.expiration_date.strftime("%d/%m/%Y") if p.expiration_date else "N/A",
            p.days_remaining,
            p.status,
            p.amount_medical_plan,  
            p.amount_nutrition_plan,
        ])
    # Calculate Totals
    total_amount = sum(float(p.amount) for p in payments)
    total_medical = sum(float(p.amount_medical_plan) for p in payments)
    total_nutrition = sum(float(p.amount_nutrition_plan) for p in payments)
    
    # Summary Row
    summary_row = ["TOTAL", "", "", total_amount, "", "", "", "", total_medical, total_nutrition]
    ws.append(summary_row)
    
    # Summary Styling
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    # Response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename=relatorio_pagamentos_{timezone.now().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response
